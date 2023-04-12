import pyperclip
import pyautogui as pag
import time
import openpyxl
import os
import glob
from pathlib import Path

def paste_translate(xx,yy):
    global ls_num
    ya=290+ls_num
    pag.click(x=1536,y=ya, duration=0.7) #鉛筆ボタン
    time.sleep(1.3) #編集画面入るまでの時間
    pag.click(1843,325,duration = 0.5) # Translation 選択
    time.sleep(0.4)
    pag.click(454,271,duration=1) # 言語リスト選択

    pag.click(xx,yy,duration=0.7) # 翻訳する言語位置選択

    pag.click(360,372,duration=1) #内容画面
    pag.hotkey('ctrl','a',duration=1)
    #pag.press('del')
    time.sleep(0.5)
    pag.hotkey('ctrl', 'v')
    pag.click(975,967,duration=0.5) #Preview
    pag.click(975,967,duration = 0.5) #save
    time.sleep(0.8)
    pag.click(781,953,duration = 0.5) #save
    time.sleep(0.9)
    pag.click(942,961,duration = 0.5) #compile
    time.sleep(3)
    pag.click(1149,952,duration = 0.5) #run
    time.sleep(3)
    #pag.click(82,442,duration = 0.5)#?
    pag.click(70,118,duration = 0.5)#戻る
    time.sleep(1)
    pag.click(57,101,duration = 0.5)#戻る
    
# save button の前後time入れること 

sheet_num = 2
file_path1 = "C:\\Users\\ウヨン\\Downloads\\介護フレーズ_0410_翻訳修正_v6.xlsx"

def select_sheet(s):
    
    start_row = 4
    lesson_name=[]
    wb= openpyxl.load_workbook(file_path1)
    ws=wb.worksheets[s]
    while start_row < 201:

        cell = ws.cell(start_row,1)

        if cell.value is not None:

            lesson_name.append(cell.value)
        start_row += 1
    return lesson_name
    wb.save(i)



wb = openpyxl.load_workbook(file_path1,data_only=True)
ws=wb.worksheets[sheet_num]
start_row = 4
start_column = 6
dif_col = 3 #바뀌는 열
eg_col = 9
vt_col = 10
id_col = 11
quote_en = []
quote_vi = []
quote_id = []
B = "B:"
A = "A:"
length = select_sheet(sheet_num)

#for iii in range(len(length)): 
iii=0
    
while not ws.cell(start_row , start_column).value is None and iii<=len(length) :
        
    #英語
    quote_en.extend([B+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,eg_col).value) , A+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,eg_col).value)])
    #ベトナム語
    quote_vi.extend([B+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,vt_col).value) , A+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,vt_col).value)])
    #インドネシア語
    quote_id.extend([B+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,id_col).value) , A+str(ws.cell(start_row,start_column).value) , str(ws.cell(start_row,id_col).value)])

    if (ws.cell(row = start_row , column = dif_col).value !=  ws.cell(row = start_row + 1 , column = dif_col).value):
        
        #英語
        
        quoteline_en = "\n".join(quote_en)
        pyperclip.copy(quoteline_en)
        ls_num = iii*33
        paste_translate(406,375)
    
        #ベトナム語
        quoteline_vi = "\n".join(quote_vi)
        pyperclip.copy(quoteline_vi)
        ls_num = iii*33
        paste_translate(418,394)
    
        # #インドネシア語
        quoteline_id = "\n".join(quote_id)
        pyperclip.copy(quoteline_id)
        ls_num = iii*33
        paste_translate(421,520)
        
        quote_en = []
        quote_vi = []
        quote_id = []
        iii+=1
    start_row += 1

            
