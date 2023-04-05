import os
import glob
from pathlib import Path
import openpyxl
import pyautogui as pag
import time
import pyperclip
from openpyxl import load_workbook

# i 파일 패스
# selectsheet(시트번수번째)

i = "C:\\Users\\ウヨン\\Downloads\\MyJT建設フレーズ_0328_v1.xlsx"

def select_sheet(s):
    start_row = 4
    lesson_name=[]
    wb= openpyxl.load_workbook(i)
    ws=wb.worksheets[s]
    while start_row < 201:

        cell = ws.cell(start_row,1)

        if cell.value is not None:

            lesson_name.append(cell.value)
        start_row += 1
    return lesson_name
    wb.save(i)
    
def add_lesson(paste_sheet_num):
    ws_value_list = select_sheet(paste_sheet_num)  
    for ws_value in ws_value_list:
        
        pag.moveTo(x=95, y=799, duration=0.5) # AddLesson clcik
        pag.click(x=95, y=799, duration=0.5)
        pag.moveTo(x=849, y=480, duration=0.5) # Lesson title 移動
        pag.doubleClick(x=849, y=480, duration=0.5) #Lesson title ダブるクリック
        pyperclip.copy(ws_value)
        pag.hotkey('ctrl', 'v')
        time.sleep(0.8)
        pag.moveTo(x=872, y=638,duration=0.5)#create 
        pag.click(x=872, y=638,duration=0.5)#create click
        time.sleep(0.3)
        
#add_lesson()  # レッスン名とレッスンを追加したときに＃消して

#--------------------------------------------------------------------------------------

#print(pag.position())
def paste_cursor(a):
    yy=290+a
    pag.click(x=1538, y=yy,duration=0.6) #y+=33 pencil icon
    pag.moveTo(x=615, y=639,duration=1) #内容真ん中画面移動
    pag.click(x=615, y=639,duration=1) #真ん中クリック
    pag.hotkey('ctrl','a',duration=1)
    pag.press('del')
    pag.hotkey('ctrl','v')
    pag.click(x=783, y=955,duration=0.5)#save button
    time.sleep(0.5)
    pag.click(x=933, y=956,duration=0.7) # compile button
    time.sleep(1.8)
    pag.click(x=1842, y=754,duration=0.7) # Theme button
    pag.click(x=116, y=297,duration=1) # My Album
    pag.click(x=826, y=713,duration=1) # Unilabs Theme 
    pag.click(x=965, y=949,duration=1) # Add to Resource to List
    pag.click(x=964, y=622,duration=1) # Theme OK
    pag.click(x=783, y=955,duration=0.6)#save button
    time.sleep(0.5)
    pag.click(x=933, y=956,duration=1) # compile button
    time.sleep(1.8)
    pag.click(x=65, y=103,duration=1) # 戻るボタン


def paste_lesson(sheet_num):
    wb= openpyxl.load_workbook(i)
    ws=wb.worksheets[sheet_num]
    iii = 1 # sheet 1번째시트부터
    ws_value_list = select_sheet(sheet_num)
    length = len(ws_value_list) # 총 몇개의 스크립트가 있는지 확인

    start_col =13
    while iii <= length:
        start_row =4
        #start_col =13
        my_list=[]
        while not ws.cell(start_row,start_col).value is None:
            my_list.append(ws.cell(start_row,start_col).value)
            start_row+=1 # 한 문장씩 한스크립트가 끝날때까지
        result = ','.join(my_list).replace(',', '\n') # 한스크립트가 끝나면 리스트안의 ,로 나뉘어진 부분을 없에고 행을 나눠서 하나의 문자열로만듬
        # iii 를 pastecursor함수에 argument로 넣고 그 숫자 만큼 밑에서 y 값을 생성
        pyperclip.copy(result)
        a = (iii-1)*33
        paste_cursor(a)
        start_row = 4
        start_col+=1
        my_list=[]
        iii+=1

add_lesson(8)
paste_lesson(8)