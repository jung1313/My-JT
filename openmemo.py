import os
import glob
from pathlib import Path
import openpyxl
import json

file_path1 = "C:\\Users\\ウヨン\\Downloads\\介護フレーズ_230302 (1).xlsx"

wb = openpyxl.load_workbook(file_path1,data_only=True)
value = []
for ws in wb:
    a = ws.title
    value.append(a)
B = len(value)
Last_sheets_num = int(B)-1 #7


i = 0
start_row = 4
start_column = 3
Lesson_number = []
values=[]
j = 4

copy_row = 4

paste_column = 13

with open('C:\jocoding\Info_JTStudio.json') as f:
    jsn = json.load(f)

JTTeacher = "JTTeacher[]"
JTStudent = "JTStudent[]"

for ws in wb.worksheets:
    Lesson_number = []
    quote = []
    start_row = 4
    avatar_row = 4
    paste_row = 15
    #paste_column = 13
    image_file = []
    image_row = 4
    grammar_row = 4
    grammar_col = 13
    png_row = 4
    paste_column = 13
    i = 0
    ii = 0
    img_col = 13
    while not ws.cell(start_row , start_column).value is None:

        if ws.cell(avatar_row,5).value == "layla" or ws.cell(avatar_row,5).value == "Layla" or ws.cell(avatar_row,5).value == "Lyla":
            quote.extend([jsn['setAVatar_1'],jsn['setAvatar_2'],JTTeacher[:10] + ws.cell(start_row,6).value + JTTeacher[10:],JTStudent[:10] + ws.cell(start_row,6).value + JTStudent[10:]])
        elif ws.cell(avatar_row,5).value == "Reiko":
            quote.extend([jsn['setAva_5'],jsn['setAva_6'],JTTeacher[:10] + ws.cell(start_row,6).value + JTTeacher[10:],JTStudent[:10] + ws.cell(start_row,6).value + JTStudent[10:]])
        elif ws.cell(avatar_row,5).value == "Ken":
            quote.extend([jsn['setAva_3'],jsn['setAva_4'],JTTeacher[:10] + ws.cell(start_row,6).value + JTTeacher[10:],JTStudent[:10] + ws.cell(start_row,6).value + JTStudent[10:]])
        
        if not ws.cell(image_row,8).value == 0:
            image_file.append(ws.cell(image_row,8).value)


        if (ws.cell(row = start_row , column = start_column).value !=  ws.cell(row = start_row + 1 , column = start_column).value):
            
            for grammar in jsn["Basic_grammar"].values():
                if grammar == "ShowPicture[]":
                    ws.cell(14,img_col).value = grammar[:12]+ image_file[ii] +grammar[12:]
                else:
                    ws.cell(grammar_row,grammar_col).value = grammar
                    grammar_row += 1
            img_col += 1
            ii = ii +1
            grammar_row = 4 
            grammar_col += 1

            for quote_line in quote:
                ws.cell(paste_row,paste_column).value = quote_line   #paste_row => basic grammar が書かれた後に 12行目から
                paste_row += 1
            paste_row = 15
            paste_column += 1
            quote =[] #reset

        image_row += 1
        avatar_row += 1 
        start_row += 1   

        
    wb.save(file_path1)    
            


print("完了しました")

