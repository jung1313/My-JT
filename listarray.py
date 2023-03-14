import os
import glob
from pathlib import Path
import openpyxl
import json
import googletrans
from googletrans import Translator

file_path1 = "C:\\Users\\ウヨン\\Desktop\\介護GA\\介護フレーズtest.xlsx"

wb = openpyxl.load_workbook(file_path1,data_only=True)

start_row = 4
start_column = 6
lang_row = 4
lang_col = 6

languages = {'ベトナム語': 'vi', 'インドネシア語': 'id', '英語': 'en'}

for ws in wb.worksheets:
    ii = 4 #英語
    kk = 9
    zz = 4 # ベトナム
    ll = 10
    mm = 4
    hh = 11
    start_row = 4
    start_column = 6
    lang_row = 4
    lang_col = 6    
    print(1)
    languages = {'ベトナム語': 'vi', 'インドネシア語': 'id', '英語': 'en'}
    jp_words=[]
    start_row = 4
    lang_row = 4
    while not ws.cell(start_row , start_column).value is None:
        #jp_words.append(ws.cell(lang_row,lang_col).value)
        jp_words.append(ws.cell(start_row,start_column).value)
        start_row = start_row+1
    
    translations = {lang: [] for lang in languages}
    translator = Translator()
    for word in jp_words:
        for lang, code in languages.items():
            result = translator.translate(word, dest=code)
            translations[lang].append(result.text)
    #print(translations)
    for key in translations:
        print(f"{key}:")
        if key == 'ベトナム語':
            for item in translations[key]:
                print(1)
                #print(f"ベトナム語 item:{item}")
                ws.cell(zz,ll).value = item
                zz = zz+1
        elif key == 'インドネシア語':
            for item in translations[key]:
                #print(f"インドネシア語 item:{item}")
                ws.cell(mm,hh).value = item
                mm = mm +1
        elif key == '英語':
            for item in translations[key]:
                #print(f"英語 item:{item}")
                ws.cell(ii,kk).value = item
                ii = ii+1
wb.save(file_path1)
